from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from models import Customers
from database import session
from sqlalchemy import func

customers = session.query(Customers).all()
count = session.query(Customers).count()
count1 = session.query(Customers).get_column.count()
print (count1)
wb = Workbook()
dest_filename = 'empty_book.xlsx'
#dsadasdsadasadsds
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
print(ws3['AA10'].value)

wb.save(filename = dest_filename)